"""The ONLY door to the workbook.

Nothing else in the system opens the .xlsx. Every accessor takes a Principal and
enforces tenant isolation before returning anything, so access control is a
property of the data layer rather than a request in the system prompt.

Note `find_order` / `find_ticket`: an unauthorised lookup and a genuinely
missing record deliberately return the SAME shape of refusal. Otherwise a
customer could enumerate other tenants' order IDs by watching which ones give
"not found" versus "access denied".
"""
from __future__ import annotations

import re
from functools import lru_cache
from typing import Iterable

import openpyxl

from app import config
from app.core import clock
from app.core.models import Account, Order, OrderStatus, Plan, Ticket
from app.core.principal import AccessDenied, Perm, Principal


def _norm_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in {"true", "yes", "1", "y"}


def _norm_str(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _rows(ws) -> Iterable[dict]:
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(it)]
    for row in it:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        yield dict(zip(header, row))


class Dataset:
    """Immutable in-memory snapshot of the supplied workbook."""

    def __init__(self) -> None:
        wb = openpyxl.load_workbook(config.WORKBOOK, data_only=True)

        # The README pins the reference time for every time-based question.
        meta = {str(r[0]).strip(): r[1] for r in
                wb["README"].iter_rows(values_only=True) if r and r[0]}
        snap = clock.parse_dt(str(meta.get("Dataset snapshot", "")).split(" Asia")[0].strip())
        if snap:
            clock.set_snapshot(snap)
        self.snapshot = clock.now()
        self.readme = {k: str(v) for k, v in meta.items() if v is not None}

        self.accounts: dict[str, Account] = {}
        for r in _rows(wb["accounts"]):
            self.accounts[r["account_id"]] = Account(
                account_id=r["account_id"],
                account_name=r["account_name"],
                plan=Plan(str(r["plan"]).strip()),
                status=str(r["status"]).strip(),
                csm=_norm_str(r.get("csm")),
                contract_file=_norm_str(r.get("contract_file")),
                premium_support=_norm_bool(r.get("premium_support")),
                notes=_norm_str(r.get("notes")),
            )

        self.orders: dict[str, Order] = {}
        for r in _rows(wb["orders"]):
            self.orders[r["order_id"]] = Order(
                order_id=r["order_id"],
                account_id=r["account_id"],
                carrier=str(r["carrier"]).strip(),
                status=OrderStatus(str(r["status"]).strip().upper()),
                booked_at=clock.parse_dt(r.get("booked_at")),
                pickup_window_start=clock.parse_dt(r.get("pickup_window_start")),
                pickup_window_end=clock.parse_dt(r.get("pickup_window_end")),
                pickup_actual_at=clock.parse_dt(r.get("pickup_actual_at")),
                shipment_fee_inr=float(r.get("shipment_fee_inr") or 0),
                carrier_fault=_norm_bool(r.get("carrier_fault")),
                customer_fault=_norm_bool(r.get("customer_fault")),
                cancellation_requested_at=clock.parse_dt(r.get("cancellation_requested_at")),
                notes=_norm_str(r.get("notes")),
            )

        self.tickets: dict[str, Ticket] = {}
        for r in _rows(wb["tickets"]):
            self.tickets[r["ticket_id"]] = Ticket(
                ticket_id=r["ticket_id"],
                account_id=r["account_id"],
                created_at=clock.parse_dt(r.get("created_at")),
                status=str(r.get("status") or "open").strip(),
                subject=_norm_str(r.get("subject")) or "",
                description=_norm_str(r.get("description")) or "",
                channel=_norm_str(r.get("channel")) or "",
                assigned_to=_norm_str(r.get("assigned_to")),
                last_customer_message_at=clock.parse_dt(r.get("last_customer_message_at")),
                historical_resolution=_norm_str(r.get("historical_resolution")),
            )

        self._load_synthetic_history()

    def _load_synthetic_history(self) -> None:
        """Optional extra ticket history for the Signal Board.

        The brief explicitly invites additional data. Seven tickets cannot
        demonstrate "a sudden increase in similar complaints", so we layer a
        clearly-flagged synthetic backlog on top. Every synthetic record carries
        synthetic=True and is excluded from any customer-facing answer, so it can
        never be mistaken for supplied ground truth.
        """
        path = config.ROOT / "data" / "synthetic_tickets.json"
        if not path.exists():
            return
        import json
        for r in json.loads(path.read_text()):
            t = Ticket(**{**r, "synthetic": True,
                          "created_at": clock.parse_dt(r.get("created_at")),
                          "last_customer_message_at": clock.parse_dt(r.get("last_customer_message_at"))})
            self.tickets[t.ticket_id] = t


@lru_cache(maxsize=1)
def dataset() -> Dataset:
    return Dataset()


# ---------------------------------------------------------------------------
# Guarded accessors -- everything below requires a Principal
# ---------------------------------------------------------------------------

class Repository:
    def __init__(self, principal: Principal):
        self.p = principal
        self.ds = dataset()

    # -- helpers ------------------------------------------------------------
    def _visible_accounts(self) -> list[Account]:
        if self.p.can(Perm.READ_ALL_ACCOUNTS):
            return list(self.ds.accounts.values())
        return [a for a in self.ds.accounts.values() if a.account_id == self.p.account_id]

    OPAQUE = ("No {kind} matching {ref!r} is available in this session. "
              "It either does not exist or is outside your access scope.")

    def _not_found(self, kind: str, ref: str, internal: str | None = None) -> AccessDenied:
        return AccessDenied(
            self.OPAQUE.format(kind=kind, ref=ref),
            resource=ref,
            internal_reason=internal or f"{kind} {ref!r} not present in dataset",
        )

    def _guard(self, kind: str, ref: str, account_id: str) -> None:
        """Convert a tenant-isolation failure into the SAME opaque error a
        missing record produces, so the two cannot be told apart from outside."""
        try:
            self.p.assert_account_access(account_id)
        except AccessDenied as e:
            raise self._not_found(kind, ref, internal=e.internal_reason) from None

    # -- accounts -----------------------------------------------------------
    def get_account(self, account_id: str) -> Account:
        acct = self.ds.accounts.get(account_id)
        if acct is None:
            raise self._not_found("account", account_id)
        self._guard("account", account_id, account_id)
        return acct

    def list_accounts(self) -> list[Account]:
        return self._visible_accounts()

    def resolve_account(self, hint: str) -> Account:
        """Accept an account id OR a company name ('Northstar')."""
        h = (hint or "").strip().lower()
        for a in self._visible_accounts():
            if a.account_id.lower() == h or a.account_name.lower() == h:
                return a
        for a in self._visible_accounts():
            if h and (h in a.account_name.lower() or a.account_name.lower().startswith(h)):
                return a
        raise self._not_found("account", hint)

    # -- orders -------------------------------------------------------------
    def get_order(self, order_id: str) -> Order:
        o = self.ds.orders.get((order_id or "").strip().upper())
        if o is None:
            raise self._not_found("order", order_id)
        self._guard("order", order_id, o.account_id)   # <-- tenant isolation
        return o

    def list_orders(self, account_id: str | None = None) -> list[Order]:
        out = []
        for o in self.ds.orders.values():
            try:
                self.p.assert_account_access(o.account_id)
            except AccessDenied:
                continue
            if account_id and o.account_id != account_id:
                continue
            out.append(o)
        return sorted(out, key=lambda x: x.order_id)

    # -- tickets ------------------------------------------------------------
    def get_ticket(self, ticket_id: str) -> Ticket:
        t = self.ds.tickets.get((ticket_id or "").strip().upper())
        if t is None:
            raise self._not_found("ticket", ticket_id)
        self._guard("ticket", ticket_id, t.account_id)
        return t

    def list_tickets(self, account_id: str | None = None, *,
                     only_open: bool = False,
                     include_synthetic: bool = True) -> list[Ticket]:
        out = []
        for t in self.ds.tickets.values():
            try:
                self.p.assert_account_access(t.account_id)
            except AccessDenied:
                continue
            if account_id and t.account_id != account_id:
                continue
            if only_open and not t.is_open:
                continue
            # Synthetic history exists to give the internal Signal Board a
            # realistic baseline. It is NOT supplied ground truth, so it must
            # never reach a customer-facing answer.
            if t.synthetic and (not include_synthetic or self.p.is_customer):
                continue
            out.append(t)
        return sorted(out, key=lambda x: (x.created_at or clock.now()), reverse=True)

    def search_tickets(self, query: str, *, limit: int = 10) -> list[Ticket]:
        terms = [w for w in re.findall(r"[a-z0-9]+", (query or "").lower()) if len(w) > 2]
        scored = []
        for t in self.list_tickets(include_synthetic=not self.p.is_customer):
            blob = f"{t.subject} {t.description}".lower()
            score = sum(blob.count(w) for w in terms)
            if score:
                scored.append((score, t))
        scored.sort(key=lambda x: (-x[0], x[1].ticket_id))
        return [t for _, t in scored[:limit]]
