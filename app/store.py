"""Reference data in SQL, loaded once from the workbook.

The .xlsx is the SOURCE, not the runtime store. It is read at build time into
SQLite so that every lookup in the product is a query with a WHERE clause
instead of a scan over parsed spreadsheet rows -- which is what makes account
scoping enforceable in one place rather than remembered at each call site.

SQLite rather than Postgres on purpose: there is no server to start, the file
sits next to the dataset it came from, and nothing about this workload wants a
second process. Swapping to Postgres later is a connection string and the same
SQL, because none of it uses SQLite-specific syntax.
"""
from __future__ import annotations

import sqlite3
import threading
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl

from app import config

DB_PATH = config.VAR_DIR / "parcelpilot.db"
_LOCK = threading.RLock()
_conn: sqlite3.Connection | None = None

SCHEMA = """
CREATE TABLE IF NOT EXISTS accounts (
  account_id      TEXT PRIMARY KEY,
  account_name    TEXT NOT NULL,
  plan            TEXT NOT NULL,
  status          TEXT,
  csm             TEXT,
  contract_file   TEXT,
  premium_support INTEGER,
  notes           TEXT
);

CREATE TABLE IF NOT EXISTS orders (
  order_id                 TEXT PRIMARY KEY,
  account_id               TEXT NOT NULL REFERENCES accounts(account_id),
  carrier                  TEXT,
  status                   TEXT,
  booked_at                TEXT,
  pickup_window_start      TEXT,
  pickup_window_end        TEXT,
  pickup_actual_at         TEXT,
  shipment_fee_inr         REAL,
  carrier_fault            INTEGER,
  customer_fault           INTEGER,
  cancellation_requested_at TEXT,
  notes                    TEXT
);
CREATE INDEX IF NOT EXISTS ix_orders_account ON orders(account_id);

CREATE TABLE IF NOT EXISTS tickets (
  ticket_id                TEXT PRIMARY KEY,
  account_id               TEXT NOT NULL REFERENCES accounts(account_id),
  created_at               TEXT,
  status                   TEXT,
  subject                  TEXT,
  description              TEXT,
  channel                  TEXT,
  assigned_to              TEXT,
  last_customer_message_at TEXT,
  -- The README warns these may be WRONG. Kept, because the product has to be
  -- able to contradict them; never treated as authority.
  historical_resolution    TEXT
);
CREATE INDEX IF NOT EXISTS ix_tickets_account ON tickets(account_id);

-- Conversations. Kept in the same file as the records they are about, so a
-- transcript and the data it was computed from can never be restored from
-- different backups.
CREATE TABLE IF NOT EXISTS chats (
  id          TEXT PRIMARY KEY,
  user_id     TEXT NOT NULL,
  user_name   TEXT NOT NULL,
  role        TEXT NOT NULL,
  account_id  TEXT,
  subject_ref TEXT,          -- the order/ticket this was opened from
  title       TEXT,
  started_at  TEXT NOT NULL,
  last_at     TEXT NOT NULL,
  turns       INTEGER NOT NULL DEFAULT 0
);
CREATE INDEX IF NOT EXISTS ix_chat_user ON chats(user_id, last_at DESC);

-- One row per exchange. `payload` is the whole Answer: the decision, the rule
-- chain, the passages and the excluded sources. That is what makes an employee
-- review a verification rather than a re-reading.
CREATE TABLE IF NOT EXISTS chat_turns (
  id              INTEGER PRIMARY KEY AUTOINCREMENT,
  conversation_id TEXT NOT NULL REFERENCES chats(id) ON DELETE CASCADE,
  seq             INTEGER NOT NULL,
  question        TEXT NOT NULL,
  answer          TEXT NOT NULL,
  payload         TEXT NOT NULL,
  created_at      TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS ix_chat_turns ON chat_turns(conversation_id, seq);
"""

_SHEETS = {
    "accounts": ("account_id", "account_name", "plan", "status", "csm",
                 "contract_file", "premium_support", "notes"),
    "orders": ("order_id", "account_id", "carrier", "status", "booked_at",
               "pickup_window_start", "pickup_window_end", "pickup_actual_at",
               "shipment_fee_inr", "carrier_fault", "customer_fault",
               "cancellation_requested_at", "notes"),
    "tickets": ("ticket_id", "account_id", "created_at", "status", "subject",
                "description", "channel", "assigned_to",
                "last_customer_message_at", "historical_resolution"),
}

_BOOL_COLS = {"premium_support", "carrier_fault", "customer_fault"}


def conn() -> sqlite3.Connection:
    global _conn
    with _LOCK:
        if _conn is None:
            _conn = sqlite3.connect(DB_PATH, check_same_thread=False)
            _conn.row_factory = sqlite3.Row
            _conn.execute("PRAGMA journal_mode=WAL")
            _conn.execute("PRAGMA foreign_keys=ON")
            _conn.executescript(SCHEMA)
            _conn.commit()
        return _conn


def rows(sql: str, args: Iterable = ()) -> list[dict]:
    with _LOCK:
        return [dict(r) for r in conn().execute(sql, tuple(args)).fetchall()]


def one(sql: str, args: Iterable = ()) -> dict | None:
    r = rows(sql, args)
    return r[0] if r else None


def _clean(col: str, v: Any) -> Any:
    if v is None:
        return None
    if col in _BOOL_COLS:
        if isinstance(v, bool):
            return int(v)
        return int(str(v).strip().lower() in {"true", "yes", "1", "y"})
    if isinstance(v, datetime):
        # Stored as ISO text so ordering and comparison work in plain SQL.
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, float) and col.endswith("_inr"):
        return v
    s = str(v).strip()
    return s or None


def load_workbook(path: Path | None = None) -> dict[str, int]:
    """Import the workbook into SQL. Idempotent -- safe to run at every boot."""
    wb = openpyxl.load_workbook(path or config.WORKBOOK, data_only=True)
    counts: dict[str, int] = {}
    with _LOCK:
        c = conn()
        # Children before parents, or the DELETE trips the foreign key that
        # exists precisely to stop an order outliving its account.
        for sheet in reversed(list(_SHEETS)):
            c.execute(f"DELETE FROM {sheet}")
        for sheet, cols in _SHEETS.items():
            ws = wb[sheet]
            it = ws.iter_rows(values_only=True)
            header = [str(h).strip() if h else "" for h in next(it)]
            idx = {name: header.index(name) for name in cols if name in header}
            missing = [name for name in cols if name not in idx]
            if missing:
                raise RuntimeError(
                    f"{sheet} sheet is missing column(s) {missing}. The workbook "
                    f"shape changed; fix the mapping rather than guessing.")
            n = 0
            for row in it:
                if row is None or all(v is None for v in row):
                    continue
                values = [_clean(name, row[idx[name]]) for name in cols]
                c.execute(f"INSERT INTO {sheet} ({','.join(cols)}) "
                          f"VALUES ({','.join('?' * len(cols))})", values)
                n += 1
            counts[sheet] = n
        c.commit()
    return counts


def snapshot() -> datetime:
    """The reference time for ALL time-based reasoning, from the README sheet.

    Never wall-clock: an SLA answer that changes because someone re-ran the demo
    on a Tuesday is not a deterministic answer.
    """
    wb = openpyxl.load_workbook(config.WORKBOOK, data_only=True)
    for row in wb["README"].iter_rows(values_only=True):
        if row and row[0] and "snapshot" in str(row[0]).lower():
            raw = str(row[1]).split("Asia/")[0].strip()
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
                try:
                    return datetime.strptime(raw, fmt).replace(tzinfo=config.TIMEZONE)
                except ValueError:
                    continue
    return datetime.strptime(config.SNAPSHOT_FALLBACK,
                             "%Y-%m-%d %H:%M").replace(tzinfo=config.TIMEZONE)


# --------------------------------------------------------------------------
# Conversations
# --------------------------------------------------------------------------

def _now() -> str:
    from datetime import datetime, timezone
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def start_conversation(user, subject_ref: str | None = None,
                       title: str | None = None) -> str:
    import uuid
    cid = uuid.uuid4().hex[:12]
    ts = _now()
    with _LOCK:
        conn().execute(
            """INSERT INTO chats
               (id, user_id, user_name, role, account_id, subject_ref, title,
                started_at, last_at)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (cid, user.user_id, user.name, user.role, user.account_id,
             subject_ref, title, ts, ts))
        conn().commit()
    return cid


def add_turn(cid: str, question: str, answer: str, payload: dict) -> None:
    import json as _json
    with _LOCK:
        c = conn()
        seq = (c.execute("SELECT COALESCE(MAX(seq),0)+1 FROM chat_turns "
                         "WHERE conversation_id=?", (cid,)).fetchone()[0])
        c.execute("""INSERT INTO chat_turns
                     (conversation_id, seq, question, answer, payload, created_at)
                     VALUES (?,?,?,?,?,?)""",
                  (cid, seq, question, answer,
                   _json.dumps(payload, default=str)[:60000], _now()))
        c.execute("""UPDATE chats
                     SET last_at = ?, turns = turns + 1,
                         title = COALESCE(NULLIF(title,''), ?)
                     WHERE id = ?""", (_now(), question[:90], cid))
        c.commit()


def conversation(cid: str) -> dict | None:
    return one("SELECT * FROM chats WHERE id = ?", (cid,))


def conversation_turns(cid: str) -> list[dict]:
    import json as _json
    out = rows("SELECT * FROM chat_turns WHERE conversation_id = ? ORDER BY seq", (cid,))
    for t in out:
        try:
            t["payload"] = _json.loads(t["payload"])
        except (TypeError, ValueError):
            t["payload"] = {}
    return out


def list_conversations(user_id: str | None = None, role: str | None = None,
                       limit: int = 50) -> list[dict]:
    where, args = ["turns > 0"], []
    if user_id:
        where.append("user_id = ?"); args.append(user_id)
    if role:
        where.append("role = ?"); args.append(role)
    args.append(limit)
    return rows(f"SELECT * FROM chats WHERE {' AND '.join(where)} "
                f"ORDER BY last_at DESC LIMIT ?", args)
